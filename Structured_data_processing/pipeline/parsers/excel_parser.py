"""Excel parser — normalizes all workbook sheets to long-form fragments.

Strategy: Read with openpyxl (data_only=True to resolve formulas),
reshape every sheet to (model_id, metric_name, value, source_file, sheet_name).

Long form is the only shape that merges cleanly when files carry
disjoint metric sets.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import openpyxl

from pipeline.config import (
    EXCEL_DATA_START_ROW,
    EXCEL_HEADER_ROW,
    SKIP_SHEETS,
    ESTATE_SHEETS,
    MIXED_SHEETS,
)
from pipeline.models import Metric

logger = logging.getLogger(__name__)


def _safe_value(val: Any) -> Any:
    """Coerce cell value to a JSON-safe type."""
    if val is None:
        return None
    if isinstance(val, (int, float, bool)):
        return val
    return str(val).strip()


def parse_excel_workbook(
    path: Path,
    *,
    domain: str,
) -> dict[str, list[Metric]]:
    """Parse a single Excel workbook into per-model metric fragments.

    Returns:
        dict mapping canonical "Model N" → list[Metric]
        (caller is responsible for ID resolution).
    """
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    filename = path.name
    fragments: dict[str, list[Metric]] = {}

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            logger.debug("Skipping overview sheet: %s/%s", filename, sheet_name)
            continue

        if sheet_name in ESTATE_SHEETS or sheet_name in MIXED_SHEETS:
            # These are handled by generic_data.py, not here
            logger.debug("Deferring estate/mixed sheet: %s/%s", filename, sheet_name)
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < EXCEL_HEADER_ROW:
            logger.warning("Sheet %s/%s has fewer than %d rows, skipping",
                           filename, sheet_name, EXCEL_HEADER_ROW)
            continue

        # Header is at row index (EXCEL_HEADER_ROW - 1)
        header_idx = EXCEL_HEADER_ROW - 1
        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[header_idx])]

        # Model column is always the first column
        model_col = 0
        metric_cols = list(range(1, len(headers)))

        source_tag = f"{filename}/{sheet_name}"

        # Data starts at EXCEL_DATA_START_ROW (1-indexed) → index = DATA_START - 1
        data_start_idx = EXCEL_DATA_START_ROW - 1
        for row_idx in range(data_start_idx, len(rows)):
            row = rows[row_idx]
            if not row or not row[model_col]:
                continue

            raw_model_id = str(row[model_col]).strip()

            if raw_model_id not in fragments:
                fragments[raw_model_id] = []

            for col_idx in metric_cols:
                if col_idx >= len(row):
                    continue
                metric_name = headers[col_idx]
                value = _safe_value(row[col_idx])

                fragments[raw_model_id].append(
                    Metric(name=metric_name, value=value, source=source_tag)
                )

    wb.close()
    logger.info("Parsed %s: %d models, %d total metrics",
                filename, len(fragments),
                sum(len(v) for v in fragments.values()))
    return fragments


def parse_quarterly_trend(path: Path) -> dict[str, list[Metric]]:
    """Parse the Quarterly_Trend sheet which has 400 rows (4 quarters × 100 models).

    Each model gets metrics tagged with their quarter for temporal context.
    """
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    fragments: dict[str, list[Metric]] = {}
    filename = path.name
    source_tag = f"{filename}/Quarterly_Trend"

    if "Quarterly_Trend" not in wb.sheetnames:
        wb.close()
        return fragments

    ws = wb["Quarterly_Trend"]
    rows = list(ws.iter_rows(values_only=True))

    header_idx = EXCEL_HEADER_ROW - 1
    if len(rows) <= header_idx:
        wb.close()
        return fragments

    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[header_idx])]

    data_start_idx = EXCEL_DATA_START_ROW - 1
    for row_idx in range(data_start_idx, len(rows)):
        row = rows[row_idx]
        if not row or not row[0]:
            continue

        raw_model_id = str(row[0]).strip()
        quarter = str(row[1]).strip() if len(row) > 1 and row[1] else "Unknown"

        if raw_model_id not in fragments:
            fragments[raw_model_id] = []

        # Metrics from col 2 onward, prefixed with quarter
        for col_idx in range(2, min(len(headers), len(row))):
            metric_name = f"{quarter}/{headers[col_idx]}"
            value = _safe_value(row[col_idx])
            fragments[raw_model_id].append(
                Metric(name=metric_name, value=value, source=source_tag)
            )

    wb.close()
    logger.info("Parsed Quarterly_Trend: %d model-quarter entries", len(fragments))
    return fragments


def parse_all_excel(excel_files: dict[str, Path]) -> dict[str, dict[str, list[Metric]]]:
    """Parse all Excel workbooks.

    Returns:
        dict mapping domain → {raw_model_id → [Metric]}
    """
    all_fragments: dict[str, dict[str, list[Metric]]] = {}

    for domain, path in excel_files.items():
        if not path.exists():
            logger.warning("Excel file not found: %s", path)
            continue

        logger.info("Parsing Excel workbook: %s (domain: %s)", path.name, domain)
        fragments = parse_excel_workbook(path, domain=domain)
        all_fragments[domain] = fragments

        # Special handling for Quarterly_Trend in backtesting workbook
        if domain == "backtesting":
            trend_fragments = parse_quarterly_trend(path)
            if trend_fragments:
                # Merge trend data into backtesting fragments
                for model_id, metrics in trend_fragments.items():
                    if model_id in all_fragments[domain]:
                        all_fragments[domain][model_id].extend(metrics)
                    else:
                        all_fragments[domain][model_id] = metrics

    return all_fragments
