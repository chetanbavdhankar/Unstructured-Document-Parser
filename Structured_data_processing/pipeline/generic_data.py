"""Generic / estate-wide data collector.

Handles data that doesn't key to a single model:
- Coverage Gaps (estate-level gap analysis)
- Sampling Methodology (per-family, not per-model)
- DQ Issue Log (mixed: some model-level, some feed-level)
- MRM Issue Tracking (mixed: some model-level, some generic)
- Generic references (research papers, etc.)
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import openpyxl

from pipeline.config import EXCEL_HEADER_ROW, EXCEL_DATA_START_ROW
from pipeline.models import EstateDossier

logger = logging.getLogger(__name__)


def _safe_value(val: Any) -> Any:
    """Coerce cell value to a JSON-safe type."""
    if val is None:
        return None
    if isinstance(val, (int, float, bool)):
        return val
    return str(val).strip()


def _read_sheet(path: Path, sheet_name: str) -> tuple[list[str], list[list[Any]]]:
    """Read a sheet from a workbook, returning headers and data rows."""
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        wb.close()
        return [], []

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < EXCEL_HEADER_ROW:
        return [], []

    header_idx = EXCEL_HEADER_ROW - 1
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[header_idx])]

    data_start_idx = EXCEL_DATA_START_ROW - 1
    data_rows = []
    for row in rows[data_start_idx:]:
        if row and any(cell is not None for cell in row):
            data_rows.append([_safe_value(c) for c in row])

    return headers, data_rows


def collect_coverage_gaps(coverage_path: Path) -> list[dict]:
    """Collect estate-level coverage gaps from Detection_Coverage.xlsx."""
    headers, data_rows = _read_sheet(coverage_path, "Coverage_Gaps")
    if not headers:
        return []

    gaps = []
    for row in data_rows:
        gap = {}
        for i, h in enumerate(headers):
            if i < len(row):
                gap[h.lower().replace(" ", "_")] = row[i]
        # Normalize key names for downstream use
        gap.setdefault("typology_segment", gap.get("typology_/_segment", ""))
        gap.setdefault("description", gap.get("gap_description", ""))
        gaps.append(gap)

    logger.info("Collected %d coverage gaps", len(gaps))
    return gaps


def collect_sampling_methodology(backtesting_path: Path) -> dict[str, str]:
    """Collect family-level sampling methodology.

    Returns dict of family_name → formatted description string.
    """
    headers, data_rows = _read_sheet(backtesting_path, "Sampling_Methodology")
    if not headers:
        return {}

    family_methods: dict[str, str] = {}
    for row in data_rows:
        if not row or not row[0]:
            continue
        family = str(row[0])
        parts = []
        for i in range(1, min(len(headers), len(row))):
            if row[i]:
                parts.append(f"{headers[i]}: {row[i]}")
        family_methods[family] = "; ".join(parts)

    logger.info("Collected sampling methodology for %d families", len(family_methods))
    return family_methods


def collect_issue_log(
    path: Path,
    sheet_name: str,
) -> tuple[list[dict], list[dict]]:
    """Collect issues from DQ_Issue_Log or Issue_Tracking sheets.

    Returns:
        (model_issues, estate_issues):
        - model_issues: dicts with a 'model' key → routed to that model's dossier
        - estate_issues: dicts without a resolvable model → estate context
    """
    headers, data_rows = _read_sheet(path, sheet_name)
    if not headers:
        return [], []

    model_issues = []
    estate_issues = []

    # Find the 'Model' column
    model_col_idx = None
    for i, h in enumerate(headers):
        if h.lower() == "model":
            model_col_idx = i
            break

    for row in data_rows:
        issue = {}
        for i, h in enumerate(headers):
            if i < len(row):
                issue[h.lower().replace(" ", "_")] = row[i]

        if model_col_idx is not None and model_col_idx < len(row) and row[model_col_idx]:
            issue["model"] = str(row[model_col_idx])
            model_issues.append(issue)
        else:
            estate_issues.append(issue)

    logger.info("Collected %d model-level and %d estate-level issues from %s",
                len(model_issues), len(estate_issues), sheet_name)
    return model_issues, estate_issues


def collect_generic_references(ref_paths: list[Path]) -> list[dict]:
    """Collect generic reference documents (research papers, etc.).

    These are not model-specific; they are estate-wide context.
    """
    refs = []
    for path in ref_paths:
        if not path.exists():
            logger.warning("Reference file not found: %s", path)
            continue

        content = path.read_text(encoding="utf-8")
        refs.append({
            "filename": path.name,
            "path": str(path),
            "content_preview": content[:500] + "..." if len(content) > 500 else content,
            "size_bytes": len(content.encode("utf-8")),
        })

    logger.info("Collected %d generic references", len(refs))
    return refs


def collect_all_generic_data(
    excel_files: dict[str, Path],
    generic_ref_paths: list[Path],
) -> tuple[EstateDossier, dict]:
    """Collect all estate-wide / generic data.

    Returns:
        (estate_dossier, routable_issues):
        - estate_dossier: contains all estate-level data
        - routable_issues: dict of sheet_name → list[dict] for model-level issues
          that should be routed to individual dossiers via merge.upsert_issues()
    """
    estate = EstateDossier()
    routable: dict[str, list[dict]] = {}

    # Coverage gaps
    if "coverage" in excel_files:
        estate.coverage_gaps = collect_coverage_gaps(excel_files["coverage"])

    # Sampling methodology
    if "backtesting" in excel_files:
        sampling = collect_sampling_methodology(excel_files["backtesting"])
        estate.sampling_methodology = [
            {"family": k, "methodology": v} for k, v in sampling.items()
        ]

    # DQ Issue Log
    if "data_quality" in excel_files:
        model_dq, estate_dq = collect_issue_log(
            excel_files["data_quality"], "DQ_Issue_Log"
        )
        routable["DQ_Issue_Log"] = model_dq
        estate.unrouted_dq_issues = estate_dq

    # MRM Issue Tracking
    if "governance" in excel_files:
        model_mrm, estate_mrm = collect_issue_log(
            excel_files["governance"], "Issue_Tracking"
        )
        routable["Issue_Tracking"] = model_mrm
        estate.unrouted_mrm_issues = estate_mrm

    # Tuning Recommendations (from backtesting)
    if "backtesting" in excel_files:
        model_tuning, _ = collect_issue_log(
            excel_files["backtesting"], "Tuning_Recommendations"
        )
        routable["Tuning_Recommendations"] = model_tuning

    # Generic references
    estate.generic_references = collect_generic_references(generic_ref_paths)

    return estate, routable
